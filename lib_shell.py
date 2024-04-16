from datetime import datetime

import matplotlib.pyplot as plt

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side

from prettytable import PrettyTable
import lib_branch as Br
import dataFormat as Df
import lib_database as Db
from lib_tree_editor import TreeEditor

DATABASE_NAME = "accountBook.db"
quit_commands = {'q!', 'Q!', 'quit', 'QUIT'}


class Shell:
    def __init__(self, root):
        self.prompt = "$[~HOME/]>> "  # 프롬프트 메시지 (현 브랜치 기반)
        self.root = root
        self.branch = self.root  # 현재 브랜치
        self.db = Db.SqlBox()

    # 유저 커맨드 실행
    def fetch(self, command):
        try:
            list_cmd = command.split()
            if list_cmd[0] in {'chdir', 'cd'} and len(list_cmd) == 2:
                self.chdir(list_cmd)
            elif list_cmd[0] in {'list', 'ls'} and len(list_cmd) == 1:
                self.listdir()
            elif list_cmd[0] in {'refer', 'rf'} and len(list_cmd) > 1:
                self.refer(list_cmd)
            elif list_cmd[0] == 'graph' and 1 <= len(list_cmd) < 3:
                self.graph(list_cmd)
            elif list_cmd[0] == 'excel' and len(list_cmd) == 2:
                self.display_excel(list_cmd)
            elif list_cmd[0] == 'help' and len(list_cmd) == 1:
                self.display_help()
            elif list_cmd[0] == 'sync' and len(list_cmd) == 1:
                self.synchronization()
            elif list_cmd[0] == 'tree' and len(list_cmd) == 1:
                self.tree()

        except Exception as e:
            print('!Error:', e)
            pass

    def synchronization(self):
        try:
            print("Starting synchronization. Please wait a moment...")
            self.db = Db.SqlBox()
            print("Synchronization completed successfully. The data has been updated.")
        except Exception as e:
            print(f"An error occurred during synchronization: {e}")
            print("Please try again, or contact support if the problem persists.")

    # 브랜치 이동
    def chdir(self, command: list):
        # 1. 타깃 브랜치 추적
        if len(command) == 1:
            command.append('')

        n, target = 0, command[1]
        if not target in self.branch.children:
            if target.isdigit():
                index = int(target) - 1
                if not (0 <= index < len(self.branch.children)):
                    print("!Error: Not valid path")
                    print("...")
                    return
                target = list(self.branch.children.keys())[index]

        while target.startswith('../'):
            n += 1
            target = target.lstrip('.').lstrip('/')

        navigating = self.branch
        for i in range(n):
            if navigating == self.root:
                break
            navigating = navigating.parent

        # 2. 타깃 브랜치로 이동
        try:
            if n == 0 and (target == '' or target == '-'):  # navigate to HOME
                self.branch = self.root
            elif n == 0 and target.split('/')[0] == 'HOME':  # navigate to an address referenced from HOME
                new_target = "/".join(target.split('/')[1::])
                navigating = Br.search_branch(self.root, new_target)
                if navigating:
                    self.branch = navigating
                else:
                    print("!Error: Not valid path")
                    print("...")
            elif n > 0 and target == '':  # navigate to an address referenced from current Branch (depth > 1)
                self.branch = navigating
            else:  # navigate to an address referenced from HOME (depth == 1)
                navigating = Br.search_branch(navigating, target)
                if navigating:
                    self.branch = navigating
                else:
                    print("!Error: Not valid path")
                    print("...")

            # 3. 프롬프트 메시지 업데이트
            self.prompt = f"$[~{self.branch.path}]>> "
        except Exception as e:
            print(e)

    # 자식 브랜치 출력
    def listdir(self):
        for i, child in enumerate(self.branch.children):
            print("{}. {}".format(i + 1, child))
        print('...')

    # 회계 장부 출력 (일별, 월별, 브랜치 별)
    def refer(self, list_cmd: list):
        if list_cmd[1] == '-d':  # 일별
            self.refer_daily(list_cmd)
        elif list_cmd[1] == '-m':  # 월별
            self.refer_monthly(list_cmd)
        elif list_cmd[1] == '-t':  # 브랜치 별
            self.refer_tree(list_cmd)

    # 일별 날짜 단위, 회계 장부 출력
    def refer_daily(self, list_cmd: list):
        # 1. SQL 쿼리 생성
        # 1.1. SQL: SELECT
        select_query = f"SELECT _Date AS DATE, _Branch, _CashFlow, _Description FROM {self.db.table_name}"

        # 1.2. SQL: WHERE
        where_query = " WHERE _Branch LIKE '{}%'".format(self.branch.path)
        begin_date, end_date = "0001-01-01", "9999-12-31"
        if len(list_cmd) == 3:
            # 날짜 변수 변환
            begin, end = list_cmd[2].split('~')
            if begin != '':
                begin_date = Df.format_date(begin)
            if end != '':
                end_date = Df.format_date(end)

            # 잘못된 날짜 입력시, 에러메시지 반환
            if (begin_date is None) or (end_date is None):
                print("!Error: NOT valid Date Input")
                print("...")
                return

        where_query += " AND (_Date BETWEEN '{}' AND '{}')".format(begin_date, end_date)

        # 1.3. SQL: ORDER
        order_query = " ORDER BY _Date;"

        # 2. SQL 실행
        sql_query = select_query + where_query + order_query
        self.db.cursor.execute(sql_query)

        # 3. 테이블 출력
        table = PrettyTable()
        table.field_names = ['DATE', 'BRANCH', 'IN', 'OUT', 'BALANCE', 'DESCRIPTION']
        count, balance = 0, 0
        total_in, total_out = 0, 0
        for row in self.db.cursor.fetchall():
            count += 1
            r_date, r_branch, r_cashflow, r_desc = row
            r_date = r_date.split()[0]

            _in = Df.format_cost(r_cashflow) if r_cashflow > 0 else '-'
            _out = Df.format_cost(-r_cashflow) if r_cashflow <= 0 else '-'

            total_in += max(r_cashflow, 0)
            total_out += -min(r_cashflow, 0)

            balance = total_in - total_out
            new_row = (r_date, r_branch, _in, _out, Df.format_cost(balance), r_desc)
            table.add_row(new_row)
        print(table)

        # 4. 요약
        print("\n*** Summary ***")
        print("- Branch: {}".format(self.branch.path))
        print("- Period: {} ~ {}".format(begin_date, end_date))
        print("- Count: {}".format(count))
        print("- Total In: {}".format(Df.format_cost(total_in)))
        print("- Total Out: {}".format(Df.format_cost(total_out)))
        print("- Balance: {}".format(Df.format_cost(balance)))
        print()

    # 월별 날짜 단위, 회계 장부 출력
    def refer_monthly(self, command: list):
        # 1. SQL 쿼리 생성
        # 1.1. SQL: SELECT
        select_query = f'''SELECT 
            STRFTIME('%Y-%m', _Date) AS MONTHLY, 
            SUM(CASE WHEN _CashFlow > 0 THEN _CashFlow ELSE 0 END) AS CASH_IN,
            SUM(CASE WHEN _CashFlow < 0 THEN _CashFlow ELSE 0 END) AS CASH_OUT
        FROM {self.db.table_name}'''

        # 1.2. SQL: WHERE - 기간 설정
        where_query = "\n   WHERE _Branch LIKE '{}%'".format(self.branch.path)
        begin_date, end_date = '0001-01-01', '9999-12-31'
        if len(command) == 3:
            begin, end = command[2].split('~')
            if begin != '':
                begin_date = Df.format_date(begin)
            if end != '':
                end_date = Df.format_date(end)
            if (begin_date is None) or (end_date is None):
                print("!Error: NOT valid Date Input")
                print("...")
                return

        where_query += " AND (_Date BETWEEN '{}' AND '{}')".format(begin_date, end_date)

        # 1.3. SQL: GROUP
        group_query = " GROUP BY MONTHLY"

        # 1.4. SQL: ORDER
        order_query = " ORDER BY MONTHLY;"

        # 2. SQL 실행
        sql_query = select_query + where_query + group_query + order_query
        self.db.cursor.execute(sql_query)

        # 3. 테이블 출력
        table = PrettyTable()
        table.field_names = ['MONTHLY', 'IN', 'OUT', 'BALANCE']
        total_in, total_out, balance = 0, 0, 0
        for row in self.db.cursor.fetchall():
            monthly, _in, _out = row
            total_in += _in
            total_out += -_out
            balance = (total_in - total_out)

            _in = Df.format_cost(_in) if _in != 0 else '-'
            _out = Df.format_cost(-_out) if _out != 0 else '-'

            new_row = (monthly, _in, _out, Df.format_cost(balance))
            table.add_row(new_row)
        print(table)
        print()

        # 4. 요약
        print("*** Summary ***")
        print("- Branch: {}".format(self.branch.path))
        print("- Period: {} ~ {}".format(begin_date, end_date))
        print("- Total In: {}".format(Df.format_cost(total_in)))
        print("- Total Out: {}".format(Df.format_cost(total_out)))
        print("- Balance: {}".format(Df.format_cost(balance)))
        print()

    # 브랜치 단위, 트리 구조 출력
    def refer_tree(self, command: list):
        # 1. 조상 노드 생성
        cost_sums = {'HOME': {'IN': 0, 'OUT': 0}}  # dictionary of each node's sum

        # 2. 기간 설정
        begin_date, end_date = '0000-01-01', '9999-12-31'

        if len(command) == 3:
            begin, end = command[2].split('~')
            if begin != '':
                begin_date = Df.format_date(begin)
            if end != '':
                end_date = Df.format_date(end)
            if (begin_date is None) or (end_date is None):
                print("!Error: NOT valid Date Input")
                print("...")
                return
        elif len(command) > 3:
            print("!Error: Too many parameters")
            print("...")
            return

        # 3. 빈 노드 채우기
        # 3.1. 모든 트랜젝션 불러오기
        sql_query = f"SELECT _Branch, _CashFlow FROM {self.db.table_name}"
        sql_query += " WHERE _Branch LIKE '{}%'".format(self.branch.path)
        sql_query += " AND '{}' <= _Date AND _Date <= '{}'".format(begin_date, end_date)

        self.db.cursor.execute(sql_query)
        res = self.db.cursor.fetchall()

        # 3.2. 기록
        for node_path, cost in res:
            cur_node = self.branch.path
            if not cur_node in cost_sums:
                cost_sums[cur_node] = {'IN': 0, 'OUT': 0}

            children_path = node_path.split(self.branch.path)[1]
            if cost > 0:
                cost_sums[cur_node]['IN'] += cost
            else:
                cost_sums[cur_node]['OUT'] -= cost

            for node in children_path.split('/'):
                if node == '':
                    continue
                cur_node += '/{}'.format(node)
                if not cur_node in cost_sums:
                    cost_sums[cur_node] = {'IN': 0, 'OUT': 0}

                if cost > 0:
                    cost_sums[cur_node]['IN'] += cost
                else:
                    cost_sums[cur_node]['OUT'] -= cost

        # 4. 재무 트리 출력
        def dfs_display(branch: Br.Branch, depth: int):
            if branch.path in cost_sums:
                total_in = cost_sums[branch.path]['IN']
                total_out = cost_sums[branch.path]['OUT']
            else:
                total_in, total_out = 0, 0
            branch_shape = '│    ' * depth + '└── '
            total_balance = Df.format_cost(total_in - total_out)
            total_in = Df.format_cost(total_in)
            total_out = Df.format_cost(total_out)
            summary = "IN:{}, OUT:{}, BAL: {}".format(total_in, total_out, total_balance)
            print("{}{}[{}]".format(branch_shape, branch.name, summary))

            for child in branch.children:
                child_branch = branch.children[child]
                dfs_display(child_branch, depth + 1)

        dfs_display(self.branch, 0)
        print("...")
        return

    # 그래프 출력
    def graph(self, list_cmd):
        if len(list_cmd) == 1:  # 수입, 지출 모두 출력
            self.graph_balance(list_cmd)
        elif list_cmd[1] == 'in':  # 수입 출력
            self.graph_in(list_cmd)
        elif list_cmd[1] == 'out':  # 지출 출력
            self.graph_out(list_cmd)

    # 수입, 지출 그래프
    def graph_balance(self, list_cmd: list):
        # 1. 브랜치 경로
        path = self.branch.path

        # 2. SQL 쿼리 생성
        # 2.1. SQL: SELECT
        select_query = f'''SELECT
            STRFTIME('%Y-%m', _Date) AS MONTHLY, 
            SUM(CASE WHEN _CashFlow > 0 THEN _CashFlow ELSE 0 END) AS CASH_IN,
            SUM(CASE WHEN _CashFlow < 0 THEN _CashFlow ELSE 0 END) AS CASH_OUT
        FROM {self.db.table_name}'''

        # 2.2. SQL: WHERE
        where_query = "\n   WHERE _Branch LIKE '{}%'".format(path)
        begin_date, end_date = '0001-01-01', '9999-12-31'
        if len(list_cmd) == 2:
            # 기간 입력 유효성 검사
            begin, end = list_cmd[2].split('~')
            if begin != '':
                begin_date = Df.format_date(begin)
            if end != '':
                end_date = Df.format_date(end)
            if (begin_date is None) or (end_date is None):
                print("!Error: NOT valid Date Input")
                print("...")
                return
        where_query += " AND (_Date BETWEEN '{}' AND '{}')".format(begin_date, end_date)

        # 2.3. SQL: GROUP
        group_query = " GROUP BY MONTHLY"

        # 2.4. SQL: ORDER
        order_query = " ORDER BY MONTHLY;"

        # 3. SQL 실행 및 데이터 조회
        sql_query = select_query + where_query + group_query + order_query

        self.db.cursor.execute(sql_query)
        total_in, total_out, balance = 0, 0, 0
        months, balances = [], []
        for row in self.db.cursor.fetchall():
            monthly, _in, _out = row
            months.append(monthly)

            total_in += _in
            total_out += (-_out)
            balances.append(total_in - total_out)

        # 4. 요약
        print()
        print("*** Summary ***")
        print("- Branch: {}".format(self.branch.path))
        print("- Period: {} ~ {}".format(begin_date, end_date))
        print("- Total In: {}".format(Df.format_cost(total_in)))
        print("- Total Out: {}".format(Df.format_cost(total_out)))
        print("- Balance: {}".format(Df.format_cost(total_in - total_out)))
        print()

        # 5. 막대그래프 출력
        plt.bar(months, balances, label='Balance')
        plt.xlabel('Month')
        plt.ylabel('Amount')
        plt.title('Monthly Balance')
        plt.legend()
        plt.show()

    # 수입 그래프
    def graph_in(self, list_cmd: list):
        # 1. 브랜치 경로
        path = self.branch.path

        # 2. SQL 쿼리 생성
        # 2.1. SQL: SELECT
        select_query = f'''SELECT 
            STRFTIME('%Y-%m', _Date) AS MONTHLY, 
            SUM(CASE WHEN _CashFlow > 0 THEN _CashFlow ELSE 0 END) AS CASH_IN
        FROM {self.db.table_name}'''

        # 2.2. SQL: WHERE
        where_query = "\n   WHERE _Branch LIKE '{}%'".format(path)
        begin_date, end_date = '0001-01-01', '9999-12-31'
        if len(list_cmd) == 3:
            # 기간 입력 유효성 검사
            begin, end = list_cmd[2].split('~')
            if begin != '':
                begin_date = Df.format_date(begin)
            if end != '':
                end_date = Df.format_date(end)
            if (begin_date is None) or (end_date is None):
                print("!Error: NOT valid Date Input")
                print("...")
                return
        where_query += " AND (_Date BETWEEN '{}' AND '{}')".format(begin_date, end_date)

        # 2.3. SQL: GROUP
        group_query = " GROUP BY MONTHLY"

        # 2.4. SQL: ORDER
        order_query = " ORDER BY MONTHLY;"

        # 3. SQL 실행
        sql_query = select_query + where_query + group_query + order_query
        self.db.cursor.execute(sql_query)
        total_in = 0
        months = []
        cash_ins = []
        for row in self.db.cursor.fetchall():
            monthly, _in = row
            months.append(monthly)
            cash_ins.append(_in)
            total_in += _in

        # 4. 요약
        print()
        print("*** Summary ***")
        print("- Branch: {}".format(self.branch.path))
        print("- Period: {} ~ {}".format(begin_date, end_date))
        print("- Total In: {}".format(Df.format_cost(total_in)))
        print()

        # 5. 막대그래프 출력
        plt.bar(months, cash_ins, label='Cash In')
        plt.xlabel('Month')
        plt.ylabel('Amount')
        plt.title('Monthly Cash Flow')
        plt.legend()
        plt.show()

    # 지출 그래프
    def graph_out(self, list_cmd: list):
        # 1. 경로
        path = self.branch.path

        # 2. SQL 쿼리 구성
        # 2.1. SQL: SELECT
        select_query = f'''SELECT 
            STRFTIME('%Y-%m', _Date) AS MONTHLY, 
            SUM(CASE WHEN _CashFlow < 0 THEN _CashFlow ELSE 0 END) AS CASH_OUT
        FROM {self.db.table_name}'''

        # 2.2. SQL: WHERE
        where_query = "\n   WHERE _Branch LIKE '{}%'".format(path)
        begin_date, end_date = '0001-01-01', '9999-12-31'
        if len(list_cmd) == 3:
            # 기간 입력 유효성 검사
            begin, end = list_cmd[2].split('~')
            if begin != '':
                begin_date = Df.format_date(begin)
            if end != '':
                end_date = Df.format_date(end)
            if (begin_date is None) or (end_date is None):
                print("!Error: NOT valid Date Input")
                print("...")
                return
        where_query += " AND (_Date BETWEEN '{}' AND '{}')".format(begin_date, end_date)

        # 2.3. SQL: GROUP
        group_query = " GROUP BY MONTHLY"

        # 2.4. SQL: ORDER
        order_query = " ORDER BY MONTHLY;"

        # 3. SQL실행
        sql_query = select_query + where_query + group_query + order_query
        self.db.cursor.execute(sql_query)
        months, cash_outs = [], []
        total_out = 0
        for row in self.db.cursor.fetchall():
            monthly, _out = row
            months.append(monthly)
            cash_outs.append(-_out)
            total_out += -_out

        # 4. 요약
        print()
        print("*** Summary ***")
        print("- Branch: {}".format(self.branch.path))
        print("- Period: {} ~ {}".format(begin_date, end_date))
        print("- Total Out: {}".format(Df.format_cost(total_out)))
        print()

        # 5. 막대그래프 출력
        plt.bar(months, cash_outs, label='Cash Out')
        plt.xlabel('Month')
        plt.ylabel('Amount')
        plt.title('Monthly Cash Flow')
        plt.legend()
        plt.show()

    # 엑셀 출력
    def display_excel(self, list_cmd: list):
        if list_cmd[1] == '-d':  # 일별 엑셀 데이터
            self.excel_daily(list_cmd)
        elif list_cmd[1] == '-m':  # 월별 엑셀 데이터
            self.excel_monthly(list_cmd)

    # 일별 엑셀 데이터 출력
    def excel_daily(self, list_cmd: list):
        # 1. SQL 쿼리 생성
        # 1.1. SQL: SELECT
        select_query = f"SELECT _Date AS _Date, _Branch, _CashFlow, _Description FROM {self.db.table_name}"

        # 1.2. SQL: WHERE
        where_query = " WHERE _Branch LIKE '{}%'".format(self.branch.path)
        begin_date, end_date = "0001-01-01", "9999-12-31"
        if len(list_cmd) == 3:
            # 날짜 변수 변환
            begin, end = list_cmd[2].split('~')
            if begin != '':
                begin_date = Df.format_date(begin)
            if end != '':
                end_date = Df.format_date(end)

            # 오입력 감지
            if (begin_date is None) or (end_date is None):
                print("!Error: NOT valid Date Input")
                print("...")
                return
        where_query += " AND (_Date BETWEEN '{}' AND '{}')".format(begin_date, end_date)

        # 1.3. SQL: ORDER
        order_query = " ORDER BY _Date;"

        # 2. SQL 실행 및 데이터 조회
        sql_query = select_query + where_query + order_query
        self.db.cursor.execute(sql_query)
        headers = ['DATE', 'BRANCH', 'IN', 'OUT', 'BALANCE', 'DESCRIPTION']
        count, balance = 0, 0
        total_in, total_out = 0, 0
        data_for_excel = []
        for row in self.db.cursor.fetchall():
            count += 1
            r_date, r_branch, r_cashflow, r_desc = row

            _in = Df.format_cost(r_cashflow) if r_cashflow > 0 else '-'
            _out = Df.format_cost(-r_cashflow) if r_cashflow <= 0 else '-'

            total_in += max(r_cashflow, 0)
            total_out += -min(r_cashflow, 0)

            balance = total_in - total_out
            new_row = (r_date, r_branch, _in, _out, Df.format_cost(balance), r_desc)
            data_for_excel.append(new_row)

        total_row = ('Total', '', Df.format_cost(total_in), Df.format_cost(total_out), Df.format_cost(balance), '')
        data_for_excel.append(total_row)

        # 3. 엑셀 파일 생성
        try:
            df = pd.DataFrame(data_for_excel, columns=headers)
            wb = Workbook()
            ws = wb.active
            width_size = [12, 12, 12, 12, 12, 12]

            # add data frame onto Excel worksheet.
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
                for c_idx, value in enumerate(row):
                    cell = ws.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                    # styling headers
                    if r_idx == 0:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center")

                    # styling total row
                    if r_idx == len(df):
                        cell.font = Font(underline="single")
                        cell.border = Border(bottom=Side(style='thin'))

                    # modify cell's width
                    width_size[c_idx] = max(len(str(value)) * 1.5, width_size[c_idx])
                    column_width = width_size[c_idx]
                    ws.column_dimensions[cell.column_letter].width = column_width

            # save file
            time = str(datetime.now().timestamp()).replace('.', '_')
            excel_filename = 'daily_{}.xlsx'.format(time)
            wb.save(excel_filename)

            # Success Message
            print(f"Report successfully saved as {excel_filename}")
            print('...')
        except Exception as e:
            print('!Error:', e)
            print('...')
            return

        # 4. 요약
        print("\n*** Summary ***")
        print("- Branch: {}".format(self.branch.path))
        print("- Period: {} ~ {}".format(begin_date, end_date))
        print("- Count: {}".format(count))
        print("- Total In: {}".format(Df.format_cost(total_in)))
        print("- Total Out: {}".format(Df.format_cost(total_out)))
        print("- Balance: {}".format(Df.format_cost(balance)))
        print()

    # 월별 엑셀 데이터 출력
    def excel_monthly(self, list_cmd: list):
        # 1. 경로
        path = self.branch.path

        # 2. SQL 쿼리 생성
        # 2.1. SQL: SELECT
        select_query = f'''SELECT 
            STRFTIME('%Y-%m', _Date) AS MONTHLY, 
            SUM(CASE WHEN _CashFlow > 0 THEN _CashFlow ELSE 0 END) AS CASH_IN,
            SUM(CASE WHEN _CashFlow < 0 THEN _CashFlow ELSE 0 END) AS CASH_OUT
        FROM {self.db.table_name}'''

        # 2.2. SQL: WHERE
        where_query = "\n   WHERE _Branch LIKE '{}%'".format(path)
        begin_date, end_date = '0001-01-01', '9999-12-31'
        if len(list_cmd) == 3:
            # 기간 입력 유효성 검사
            begin, end = list_cmd[2].split('~')
            if begin != '':
                begin_date = Df.format_date(begin)
            if end != '':
                end_date = Df.format_date(end)
            if (begin_date is None) or (end_date is None):
                print("!Error: NOT valid Date Input")
                print("...")
                return

        where_query += " AND (_Date BETWEEN '{}' AND '{}')".format(begin_date, end_date)

        # 2.3. SQL: GROUP
        group_query = " GROUP BY MONTHLY"

        # 2.4. SQL: ORDER
        order_query = " ORDER BY MONTHLY;"

        # 3. SQL 실행 및 데이터 조회
        sql_query = select_query + where_query + group_query + order_query
        self.db.cursor.execute(sql_query)

        headers = ['MONTHLY', 'IN', 'OUT', 'BALANCE']
        data_for_excel = []
        total_in, total_out, balance = 0, 0, 0
        for row in self.db.cursor.fetchall():
            monthly, _in, _out = row
            total_in += _in
            total_out += -_out
            balance = (total_in - total_out)

            _in = Df.format_cost(_in) if _in != 0 else '-'
            _out = Df.format_cost(-_out) if _out != 0 else '-'

            new_row = (monthly, _in, _out, Df.format_cost(balance))
            data_for_excel.append(new_row)

        total_row = ('Total', Df.format_cost(total_in), Df.format_cost(total_out), Df.format_cost(balance))
        data_for_excel.append(total_row)

        # 5. 엑셀 파일 생성
        try:
            df = pd.DataFrame(data_for_excel, columns=headers)
            wb = Workbook()
            ws = wb.active
            width_size = [12, 12, 12, 12]

            # add data frame onto Excel worksheet.
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
                for c_idx, value in enumerate(row):
                    cell = ws.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                    # styling headers
                    if r_idx == 0:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center")

                    # styling total row
                    if r_idx == len(df):
                        cell.font = Font(underline="single")
                        cell.border = Border(bottom=Side(style='thin'))

                    # modify cell's width
                    width_size[c_idx] = max(len(str(value)) * 1.5, width_size[c_idx])
                    column_width = width_size[c_idx]
                    ws.column_dimensions[cell.column_letter].width = column_width

            # save file
            time = str(datetime.now().timestamp()).replace('.', '_')
            excel_filename = 'monthly_{}.xlsx'.format(time)
            wb.save(excel_filename)

            # Success Message
            print(f"Report successfully saved as {excel_filename}")
            print('...')
        except Exception as e:
            print('!Error:', e)
            print('...')
            return

        # 6. 요약
        print()
        print("*** Summary ***")
        print("- Branch: {}".format(self.branch.path))
        print("- Period: {} ~ {}".format(begin_date, end_date))
        print("- Total In: {}".format(Df.format_cost(total_in)))
        print("- Total Out: {}".format(Df.format_cost(total_out)))
        print("- Balance: {}".format(Df.format_cost(balance)))
        print()

    # Help 명령어
    def display_help(self):
        print("사용 가능한 명령어 및 옵션 리스트:\n")
        print("help")
        print("  도움말을 보여줍니다.")
        print()
        print("chdir, cd [브랜치명/번호] 또는 '../'")
        print("  브랜치를 이동합니다.")
        print("  예시: 'cd finance', 'cd 2', 'cd ../'")
        print()
        print("list, ls")
        print("  현재 브랜치의 자식 브랜치들을 나열합니다.")
        print()
        print("refer, rf [-d | -m | -t] [날짜범위]")
        print("  회계 장부를 참조합니다. 옵션에 따라 다양한 형태로 출력됩니다.")
        print("  -d: 일별 데이터를 보여줍니다. 예시: 'rf -d 2023-01-01~2023-01-31'")
        print("  -m: 월별 데이터를 보여줍니다. 예시: 'rf -m 2023-01'")
        print("  -t: 트리 구조로 데이터를 보여줍니다. 예시: 'rf -t'")
        print()
        print("graph [in | out] [날짜범위]")
        print("  그래프 형태로 재정 데이터를 보여줍니다. 옵션에 따라 수입, 지출, 잔액을 볼 수 있습니다.")
        print("  'graph in': 수입 데이터만 보여줍니다. 예시: 'graph in 2023-01'")
        print("  'graph out': 지출 데이터만 보여줍니다. 예시: 'graph out 2023-01'")
        print("  'graph': 수입과 지출의 잔액을 보여줍니다.")
        print()
        print("excel [-d | -m] [날짜범위]")
        print("  데이터를 Excel 파일로 출력합니다. 옵션에 따라 일별, 월별 데이터를 저장할 수 있습니다.")
        print("  -d: 일별 데이터를 Excel로 출력합니다. 예시: 'excel -d 2023-01-01~2023-01-31'")
        print("  -m: 월별 데이터를 Excel로 출력합니다. 예시: 'excel -m 2023-01'")
        print()
        print("sync")
        print("  외부에서 수정된 엑셀 파일의 내용을 프로그램에 동기화합니다. 데이터베이스가 최신 상태로 갱신됩니다.")
        print()
        print("tree")
        print("  현재 브랜치의 트리 구조를 보여줍니다.")
        print("  이 명령어는 현재 활성화된 브랜치의 전체 트리 구조를 시각적으로 표시하며,")
        print("  사용자는 'add', 'delete' 명령어를 사용하여 트리 내에서 직접 브랜치를 추가하거나 삭제할 수 있습니다.")
        print("  각 브랜치의 연결 상태와 계층을 쉽게 파악하고 관리할 수 있어 데이터 구조의 조정과 관리가 용이합니다.")
        print()
        print("q!, Q!, quit, QUIT")
        print("  프로그램을 종료합니다.")
        print()

    def tree(self):
        if self.branch.path != self.root.path:
            print('!Error:', "Home 디렉토리에서 실행해야 합니다")
        else:
            TreeEditor()


# pyinstaller --add-data "directory.json;." --icon=icon_financetree.ico -n FinanceTree3 main.py

# pyinstaller --add-data "directory.json;." --hidden-import=prettytable -n FinanceTree main.py

# pyinstaller --add-data "directory.json;." --hidden-import=prettytable --icon=icon_financetree.ico -n FinanceTree3 main.py

# pyinstaller --add-data "Tree.xlsx;data" --add-data "Transactions.xlsx;data" --icon=icon_financetree.ico -n FinanceTreeApp main.py

# pyinstaller --add-data "Tree.xlsx;." --add-data "Transactions.xlsx;." --icon=icon_financetree.ico -n FinanceTreeApp main.py

# pyinstaller --add-data "BudgetTree.json;." --add-data "Transactions.xlsx;." --icon=icon_financetree.ico -n FinanceTreeApp2 main.py
