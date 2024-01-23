from datetime import datetime

import matplotlib.pyplot as plt
from collections import deque

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side

from prettytable import PrettyTable
import branchController
import branchController as Bc
import dataFormat as Df
import branchController as Br
import sqlController

DATABASE_NAME = "accountBook.db"
quit_commands = {'q!', 'Q!'}


class Shell:
    def __init__(self):
        self.prompt = "$Finance_Tree[HOME]>> "  # prompt Message, based on current branch
        self.sqlBox = sqlController.SqlBox(DATABASE_NAME)  # from SQL Controller
        self.root = Br.make_tree()  # root Branch
        self.branch = self.root  # current Branch

    # Execute on user's command
    def fetch(self, command):
        try:
            command = command.split()
            if command[0] in {'mkdir', 'md'} and len(command) == 2:
                self.mkdir(command[1])
            elif command[0] in {'rmdir', 'rd'} and len(command) == 2:
                self.rmdir(command[1])
            elif command[0] in {'chdir', 'cd'} and len(command) == 2:
                self.navigate_dir(command)
            elif command[0] in {'list', 'ls'} and len(command) == 1:
                self.listdir()
            elif command[0] in {'refer', 'rf'} and len(command) > 1:
                self.refer(command)
            elif command[0] in {'insert', 'in'} and len(command) == 1:
                self.insert_row()
            elif command[0] in {'delete', 'del'} and len(command) <= 2:
                self.delete_row(command)
            elif command[0].lower() == 'sql':
                self.sql_execute()
            elif command[0] == 'help' and len(command) == 1:
                self.display_help()
            elif command[0] == 'mv' and len(command) == 3:
                self.move_dir(command)
            elif command[0] == 'excel' and 1 < len(command) <= 3:
                self.display_excel(command)
            elif command[0] == 'graph' and 1 <= len(command) < 3:
                self.graph(command)
        except Exception as e:
            print(e)
            pass

    # Execution Functions based on User's command.
    # Create new directory into current path
    def mkdir(self, dir_name: str):
        if dir_name.lower() == "home":
            print("!Error: Forbidden directory name('HOME').".format(dir_name))
            print('...')
        elif not set(dir_name).isdisjoint(set('\/:*?"<>|')):
            print('!Error: \/:*?"<>| cannot be included'.format(dir_name))
            print('...')
        elif dir_name in self.branch.children:
            print("!Error: Already Exist({}).".format(dir_name))
            print('...')
        else:
            Br.add_branch(self.branch, dir_name)
            print('"Directory \'{}\' has been successfully added."'.format(dir_name))
            print('...')

    # Remove a directory from current path
    def rmdir(self, dir_name: str):
        if not dir_name in self.branch.children:
            if dir_name.isdigit():
                index = int(dir_name) - 1
                dir_name = list(self.branch.children.keys())[index]
            else:
                # Fail Message
                print("!Error: There's not '{}' in this path.".format(dir_name))
                print('...')
                return

        # Ask the user for confirmation to proceed with deletion
        print("*** Warning ***")
        print('"If you delete the {}, all directories and financial transaction'.format(dir_name))
        print('details (expenses and income) will be permanently removed."\n')
        answer = input("*** Do you still wish to proceed with the deletion?(Y/N) ")
        if answer in quit_commands:
            print('User: "QUIT"')
            print('...')
            return

        if answer == 'Y' or answer == 'y' or answer.lower() == 'yes':
            print('User: "YES"\n')

            # Delete branch from Tree
            Br.delete_branch(self.branch, dir_name)
            sql_query = "DELETE FROM ACCOUNTBOOK WHERE BRANCH LIKE '{}%';".format(
                self.branch.cur_path + '/' + dir_name)
            self.sqlBox.cursor.execute(sql_query)
            self.sqlBox.database.commit()

            # Success Message
            print('"Directory \'{}\' has been successfully deleted."'.format(dir_name))
            print('...')
        else:
            print('User: "NO"')
            print('...')
            return

    # Navigate a directory based on user input
    def navigate_dir(self, command: list):
        # 1. Get path's depth and target
        if len(command) == 1:
            command.append('')
        n, target = 0, command[1]
        if not target in self.branch.children:
            if target.isdigit():
                index = int(target) - 1
                if not (0 <= index < len(self.branch.children)):
                    print("!Error: Not valid path")
                    print("...")
                    return
                target = list(self.branch.children.keys())[index]
        while target.startswith('../'):
            n += 1
            target = target.lstrip('.').lstrip('/')

        # 2. Get target Branch
        navigating = self.branch
        for i in range(n):
            if navigating == self.root:
                break
            navigating = navigating.parent

        # 3. Navigate to target Branch
        try:
            if n == 0 and (target == '' or target == '-'):  # navigate to HOME
                self.branch = self.root
            elif n == 0 and target.split('/')[0] == 'HOME':  # navigate to an address referenced from HOME
                new_target = "/".join(target.split('/')[1::])
                navigating = Br.search_branch(self.root, new_target)
                if navigating:
                    self.branch = navigating
                else:
                    print("!Error: Not valid path")
                    print("...")
            elif n > 0 and target == '':  # navigate to an address referenced from current Branch (depth > 1)
                self.branch = navigating
            else:  # navigate to an address referenced from HOME (depth == 1)
                navigating = Br.search_branch(navigating, target)
                if navigating:
                    self.branch = navigating
                else:
                    print("!Error: Not valid path")
                    print("...")

            # update the message in the prompt window
            self.prompt = "$Finance_Tree[{}]>> ".format(self.branch.cur_path)
        except Exception as e:
            print(e)

    # List Directory
    def listdir(self):
        for i, child in enumerate(self.branch.children):
            print("{}. {}".format(i + 1, child))
        print('...')

    # Refer data's from Database
    def refer(self, command: list):
        if command[1] == '-d':  # refer daily
            self.refer_daily(command)
        elif command[1] == '-m':  # refer monthly
            self.refer_monthly(command)
        elif command[1] == '-t':  # refer tree structure
            self.refer_tree(command)

    # Creating a table containing data aggregated by date.
    def refer_daily(self, command: list):
        # 1. get Query
        # 1.1. get SELECT query
        select_query = "SELECT TRANSACTION_DATE AS DATE, BRANCH, CASHFLOW, DESCRIPTION FROM ACCOUNTBOOK"

        # 1.2. get WHERE query
        where_query = " WHERE BRANCH LIKE '{}%'".format(self.branch.cur_path)
        begin_date, end_date = "0001-01-01", "9999-12-31"
        if len(command) == 3:
            # convert Date value
            begin, end = command[2].split('~')
            if begin != '':
                begin_date = Df.format_date(begin)
            if end != '':
                end_date = Df.format_date(end)

            # if invalid input date, return Error
            if (begin_date is None) or (end_date is None):
                print("!Error: NOT valid Date Input")
                print("...")
                return

        where_query += " AND (TRANSACTION_DATE BETWEEN '{}' AND '{}')".format(begin_date, end_date)

        # 1.3. get ORDER query
        order_query = " ORDER BY TRANSACTION_DATE;"

        # 2. Execution on SQL
        sql_query = select_query + where_query + order_query
        self.sqlBox.cursor.execute(sql_query)

        # 3. Print table
        table = PrettyTable()
        table.field_names = ['DATE', 'BRANCH', 'IN', 'OUT', 'BALANCE', 'DESCRIPTION']
        count, balance = 0, 0
        total_in, total_out = 0, 0
        for row in self.sqlBox.cursor.fetchall():
            count += 1
            r_date, r_branch, r_cashflow, r_desc = row

            _in = Df.format_cost(r_cashflow) if r_cashflow > 0 else '-'
            _out = Df.format_cost(-r_cashflow) if r_cashflow <= 0 else '-'

            total_in += max(r_cashflow, 0)
            total_out += -min(r_cashflow, 0)

            balance = total_in - total_out
            new_row = (r_date, r_branch, _in, _out, Df.format_cost(balance), r_desc)
            table.add_row(new_row)
        print(table)

        # 4. Print Summary
        print("\n*** Summary ***")
        print("- Branch: {}".format(self.branch.cur_path))
        print("- Period: {} ~ {}".format(begin_date, end_date))
        print("- Count: {}".format(count))
        print("- Total In: {}".format(Df.format_cost(total_in)))
        print("- Total Out: {}".format(Df.format_cost(total_out)))
        print("- Balance: {}".format(Df.format_cost(balance)))
        print()

    # Creating a table containing data aggregated by month.
    def refer_monthly(self, command: list):
        # 1. Get Path
        path = self.branch.cur_path

        # 2. get Query
        # 2.1. get SELECT query
        select_query = '''SELECT 
            STRFTIME('%Y-%m', TRANSACTION_DATE) AS MONTHLY, 
            SUM(CASE WHEN CASHFLOW > 0 THEN CASHFLOW ELSE 0 END) AS CASH_IN,
            SUM(CASE WHEN CASHFLOW < 0 THEN CASHFLOW ELSE 0 END) AS CASH_OUT
        FROM ACCOUNTBOOK'''

        # 2.2. get WHERE query
        where_query = "\n   WHERE BRANCH LIKE '{}%'".format(path)
        begin_date, end_date = '0001-01-01', '9999-12-31'
        if len(command) == 3:
            # check validity of period
            begin, end = command[2].split('~')
            if begin != '':
                begin_date = Df.format_date(begin)
            if end != '':
                end_date = Df.format_date(end)
            if (begin_date is None) or (end_date is None):
                print("!Error: NOT valid Date Input")
                print("...")
                return

        where_query += " AND (TRANSACTION_DATE BETWEEN '{}' AND '{}')".format(begin_date, end_date)

        # 2.3. get GROUP query
        group_query = " GROUP BY MONTHLY"

        # 2.4. get ORDER query
        order_query = " ORDER BY MONTHLY;"

        # 3. Execution on SQL
        sql_query = select_query + where_query + group_query + order_query
        self.sqlBox.cursor.execute(sql_query)

        # 4. Print table
        table = PrettyTable()
        table.field_names = ['MONTHLY', 'IN', 'OUT', 'BALANCE']
        total_in, total_out, balance = 0, 0, 0
        for row in self.sqlBox.cursor.fetchall():
            monthly, _in, _out = row
            total_in += _in
            total_out += -_out
            balance = (total_in - total_out)

            _in = Df.format_cost(_in) if _in != 0 else '-'
            _out = Df.format_cost(-_out) if _out != 0 else '-'

            new_row = (monthly, _in, _out, Df.format_cost(balance))
            table.add_row(new_row)
        print(table)
        print()
        print("*** Summary ***")
        print("- Branch: {}".format(self.branch.cur_path))
        print("- Period: {} ~ {}".format(begin_date, end_date))
        print("- Total In: {}".format(Df.format_cost(total_in)))
        print("- Total Out: {}".format(Df.format_cost(total_out)))
        print("- Balance: {}".format(Df.format_cost(balance)))
        print()

    # Expand the tree structure and aggregate by each branch.
    def refer_tree(self, command: list):
        # 1. Get Path
        cost_sums = {'HOME': {'IN': 0, 'OUT': 0}}  # dictionary of each node's sum
        # 2. Setting Period
        begin_date, end_date = '0000-01-01', '9999-12-31'

        if len(command) == 3:
            begin, end = command[2].split('~')
            if begin != '':
                begin_date = Df.format_date(begin)
            if end != '':
                end_date = Df.format_date(end)
            if (begin_date is None) or (end_date is None):
                print("!Error: NOT valid Date Input")
                print("...")
                return
        elif len(command) > 3:
            print("!Error: Too many parameters")
            print("...")
            return

        # 3. Fill Out Tree
        path = self.branch.cur_path
        sql_query = "SELECT BRANCH, CASHFLOW FROM ACCOUNTBOOK"
        sql_query += " WHERE BRANCH LIKE '{}%'".format(path)
        sql_query += " AND '{}' <= TRANSACTION_DATE AND TRANSACTION_DATE <= '{}'".format(begin_date, end_date)
        self.sqlBox.cursor.execute(sql_query)
        res = self.sqlBox.cursor.fetchall()

        # Recording
        for node_path, cost in res:
            cur_node = path
            if not cur_node in cost_sums:
                cost_sums[cur_node] = {'IN': 0, 'OUT': 0}

            children_path = node_path.split(path)[1]
            if cost > 0:
                cost_sums[cur_node]['IN'] += cost
            else:
                cost_sums[cur_node]['OUT'] -= cost

            for node in children_path.split('/'):
                if node == '':
                    continue
                cur_node += '/{}'.format(node)
                if not cur_node in cost_sums:
                    cost_sums[cur_node] = {'IN': 0, 'OUT': 0}

                if cost > 0:
                    cost_sums[cur_node]['IN'] += cost
                else:
                    cost_sums[cur_node]['OUT'] -= cost

        # 4. Display each Branch's Summary
        def dfs_display(branch: branchController.Branch, depth: int):
            if branch.cur_path in cost_sums:
                total_in = cost_sums[branch.cur_path]['IN']
                total_out = cost_sums[branch.cur_path]['OUT']
            else:
                total_in, total_out = 0, 0
            branch_shape = '|    ' * depth + '|-- '
            total_balance = Df.format_cost(total_in - total_out)
            total_in = Df.format_cost(total_in)
            total_out = Df.format_cost(total_out)
            summary = "IN:{}, OUT:{}, BAL: {}".format(total_in, total_out, total_balance)
            print("{}{}[{}]".format(branch_shape, branch.name, summary))

            for child in branch.children:
                child_branch = branch.children[child]
                dfs_display(child_branch, depth + 1)

        dfs_display(self.branch, 0)
        print("...")
        return

    # Insert a row into Database
    def insert_row(self):
        # 0. variables
        row = {}

        # 1. Get DATE
        user_input = input("*** Input Date(type: DATE): ")
        if user_input in quit_commands:
            print('User: "QUIT"')
            print('...')
            return

        date_str = Df.format_date(user_input)
        if date_str is None:
            print("\n!Error: Not a valid input type DATE. ({})\n...".format(user_input))
        row['TRANSACTION_DATE'] = "'{}'".format(date_str)
        week_day = Df.day_of_week(date_str)

        # 2. Get BRANCH
        row['BRANCH'] = "'{}'".format(self.branch.cur_path)

        # 3. Get CASHFLOW
        user_input = input("*** Input Cash Flow(type: INT): ")
        if user_input in quit_commands:
            print('User: "QUIT"')
            print('...')
            return

        try:
            input_cost = int(user_input)
            row['CASHFLOW'] = str(input_cost)
        except Exception as e:
            print("\n!Error: Not a valid input type INT ({}).\n...".format(input_cost))
            return

        # 4. Get DESCRIPTION
        user_input = input("*** Input Description(type: STR): ")
        if user_input in quit_commands:
            print('User: "QUIT"')
            print('...')
            return
        row['DESCRIPTION'] = "'{}'".format(user_input)

        # 5. Ask the user for confirmation to proceed with insertion
        print(' ')
        print("*** User input:")
        for header in row:
            if header == 'CASHFLOW':
                value = Df.format_cost(int(row[header]))
            elif header == 'TRANSACTION_DATE':
                value = '{}({})'.format(row[header], week_day)
            else:
                value = row[header]
            print("- {}: {}".format(header, value))
        answer = input("*** Do you want to insert following data?(Y/N) ")
        if answer in quit_commands:
            print('User: "QUIT"')
            print('...')
            return

        # 6. Execution of Insertion based on SQL
        if answer == 'Y' or answer == 'y' or answer.lower() == 'yes':
            print('User: "YES"\n')

            # create sql query
            sql_query = "INSERT INTO ACCOUNTBOOK"
            sql_query += " ({})".format(",".join(row.keys()))
            sql_query += " VALUES({})".format(",".join(row.values()))
            print('Execution on SQL \n"{}"'.format(sql_query))
            print(" ")
            try:
                self.sqlBox.cursor.execute(sql_query)
                self.sqlBox.database.commit()

                # Success Message
                print('"Successfully inserted the data"')
                print("...")
            except Exception as e:
                # Fail Message
                print("Failed to insert the data")
                print("!Error:", e)
                print("...")
        else:
            print('User: "NO"\n...')

    # Delete a row from Database
    def delete_row(self, command):
        # 1. List the deletable data in the current directory.
        # 1.1. Bring data from DATABASE, based on SQL
        sql_query = "SELECT"
        sql_query += (" ID AS NUMBER, CREATED_TIME AS UPDATED_DATE, TRANSACTION_DATE AS DATE, BRANCH, CASHFLOW, "
                      "DESCRIPTION")
        sql_query += " FROM ACCOUNTBOOK"
        sql_query += " WHERE BRANCH = '{}'".format(self.branch.cur_path)
        if len(command) == 2:
            if command[1] == '-d':
                sql_query += " ORDER BY DATE ASC, UPDATED_DATE ASC"
            else:
                print('...')
                print('!Error: {} is not valid command'.format(command[1]))
                return
        else:
            sql_query += " ORDER BY UPDATED_DATE ASC, DATE ASC"
        self.sqlBox.cursor.execute(sql_query)

        # 1.2. Create a table to show deletable datas.
        table = PrettyTable()
        table.field_names = [header[0] for header in self.sqlBox.cursor.description]
        id_box = []
        for row in self.sqlBox.cursor.fetchall():
            id_number, updated, date, branch, cashflow, description = row
            id_box.append(id_number)

            number = len(id_box)
            table.add_row(
                (number, updated, date, branch, Df.format_cost(cashflow), description)
            )

        print(table)
        table_size = len(id_box)
        if table_size == 0:
            print("\nThere is no data to delete in this directory.\n...")
            return
        elif table_size == 1:
            range_txt = ""
        else:
            range_txt = "({}~{})".format(1, table_size)

        # 2. Get User selection about the row they wish to delete
        user_input = input("*** Input number of row you want to delete{}: ".format(range_txt))
        if user_input in quit_commands:
            print('User: "QUIT"')
            print('...')
            return

        # Check if input data is number.
        try:
            del_number = int(user_input)
        except Exception as e:
            print("\n!Error: '{}' is not a number.\n...".format(user_input))
            return

        # 3. Execution of deletion based on SQL
        if del_number == table_size + 1:
            # 3.1. Delete all data in the current directory path.
            answer = input("*** Are you sure you want to delete all {} rows in this dir?(Y/N) ".format(table_size))
            if answer in quit_commands:
                print('User: "QUIT"')
                print('...')
                return

            if answer == 'Y' or answer == 'y' or answer.lower() == 'yes':
                print('User: "YES"\n')

                # Proceed deletion
                sql_query = "DELETE FROM ACCOUNTBOOK WHERE BRANCH='{}'".format(self.branch.cur_path)
                self.sqlBox.cursor.execute(sql_query)
                self.sqlBox.database.commit()
                print('*** Execution on SQL "{}" ***'.format(sql_query))

                # Success Message
                print()
                print('"Successfully deleted the data"')
            else:
                # Fail Message
                print('User: "NO"')
            print('...')
        elif not (1 <= del_number <= len(id_box)):
            # 3.2. Fail: input data is out of range.
            print("\n!Error: The ID number you entered has been out of range.")
            print("...")
        else:
            # 3.3. Delete the data corresponding to the user-specified index value.
            print(".\n.\n.\n<< Delete Row(User Selection) >>")
            print(table[del_number - 1])

            # Ask the user for confirmation to proceed with deletion
            answer = input("*** Do you really want to delete this row?(Y/N) ")
            if answer in quit_commands:
                print('User: "QUIT"')
                print('...')
                return

            if answer == 'Y' or answer == 'y' or answer.lower() == 'yes':
                print('User: "YES"')

                # create SQL
                del_id = id_box[del_number - 1]
                sql_query = "DELETE FROM ACCOUNTBOOK WHERE ID={}".format(del_id)

                try:
                    # proceed deletion on SQL
                    print('Execution on SQL "{}"'.format(sql_query))
                    print(".")
                    self.sqlBox.cursor.execute(sql_query)
                    self.sqlBox.database.commit()

                    # Success Message
                    print('"Successfully deleted the data"')
                    print('...')
                except Exception as e:
                    print("Failed to delete the data")
                    print("Error:", e)
                    print('...')
            else:
                print('User: "NO"')

    # Display Help message
    def display_help(self):
        print("Finance Tree Command Line Interface Help")
        print("---------------------------------------")
        print("Commands:")
        print("  mkdir, md <dir_name>       : Create a new directory.")
        print("  rmdir, rd <dir_name>       : Remove an existing directory.")
        print("  rmdir, rd <dir_number>     : Remove an existing directory(based on index number).")
        print("  chdir, cd <path>           : Change the current directory.")
        print("  chdir, cd <dir_number>     : Change the current directory(based on index number).")
        print("  list, ls                   : List all directories in the current directory.")
        print("  refer, rf -d <period>      : Refer daily transactions. Optionally specify a period (e.g., 2023/01/01~2023/01/31).")
        print("  refer, rf -m <period>      : Refer monthly transactions. Optionally specify a period. (e.g., 2023/01/01~2023/01/31).")
        print("  refer, rf -t               : Display the tree structure of transactions.")
        print("  insert                     : Insert a new transaction record.")
        print("  delete                     : Delete an existing transaction record(Recently).")
        print("  delete -d                  : Delete an existing transaction record(Old).")
        print("  mv <old_name> <new_name>   : Rename a directory from <old_name> to <new_name>.")
        print("  excel -d                   : Export daily transaction data to an Excel file.")
        print("  excel -m                   : Export monthly transaction data to an Excel file.")
        print("  graph                      : Display a graph of monthly balance.")
        print("  graph in                   : Display a graph of monthly cash inflow.")
        print("  graph out                  : Display a graph of monthly cash outflow.")
        print("  sql                        : Directly execute SQL queries. Only SELECT queries are allowed.")
        print("  help                       : Display this help message.")
        print()
        print("Note: Use 'q!' to quit or cancel an operation at any time.")
        print("---------------------------------------")

    # Version 1.0.2
    def sql_execute(self):
        while True:
            sql_query = input('[Sql Query]>> ')
            if len(sql_query.split()) == 0:
                continue

            if sql_query in quit_commands:
                break

            first_word = sql_query.strip().split()[0].lower()
            if first_word != 'select':
                print('Only Select')

            try:
                self.sqlBox.cursor.execute(sql_query)
                res = self.sqlBox.cursor.fetchall()

                table = PrettyTable()
                table.field_names = [header[0] for header in self.sqlBox.cursor.description]
                for row in res:
                    table.add_row(row)
                print(table)
            except Exception as e:
                print('!Error:', e)
            print('...')

    def move_dir(self, command):
        # 1. Check if exist the old one.
        old_name, new_name = command[1:]
        if not old_name in self.branch.children:
            print()
            print("!Error: There s no '{}'".format(old_name))
            return

        # 2. Check if exist the new one.
        if new_name in self.branch.children:
            print()
            print("!Error: Already exist '{}'".format(old_name))
            return

        # 3. update branch class
        self.branch.children[new_name] = self.branch.children[old_name]
        self.branch.children.pop(old_name)
        node = self.branch.children[new_name]
        node.name = new_name
        stack = deque([node])
        while stack:
            node = stack.pop()
            node.cur_path = node.parent.cur_path + '/' + node.name
            for child in node.children:
                stack.append(node.children[child])

        # 4. update json file
        Bc.modifyBranchName(self.branch, old_name, new_name)

        # 5. update Database table
        old_path = self.branch.cur_path + '/' + old_name
        new_path = self.branch.cur_path + '/' + new_name
        sql_query = 'UPDATE ACCOUNTBOOK'
        sql_query += " SET BRANCH = REPLACE(BRANCH, '{}', '{}')".format(old_path, new_path)
        sql_query += " WHERE BRANCH LIKE '{}%';".format(old_path)
        print()
        try:
            self.sqlBox.cursor.execute(sql_query)
            print('"Succesfully changed dir name: ("{}" -> "{}")"'.format(old_name, new_name))
        except Exception as e:
            print('!Error: {}'.format(e))

    def display_excel(self, command: list):
        if command[1] == '-d':
            self.excel_daily(command)
        elif command[1] == '-m':
            self.excel_monthly(command)

    def excel_daily(self, command: list):
        # 1. get Query
        # 1.1. get SELECT query
        select_query = "SELECT TRANSACTION_DATE AS DATE, BRANCH, CASHFLOW, DESCRIPTION FROM ACCOUNTBOOK"

        # 1.2. get WHERE query
        where_query = " WHERE BRANCH LIKE '{}%'".format(self.branch.cur_path)
        begin_date, end_date = "0001-01-01", "9999-12-31"
        if len(command) == 3:
            # convert Date value
            begin, end = command[2].split('~')
            if begin != '':
                begin_date = Df.format_date(begin)
            if end != '':
                end_date = Df.format_date(end)

            # if invalid input date, return Error
            if (begin_date is None) or (end_date is None):
                print("!Error: NOT valid Date Input")
                print("...")
                return
        where_query += " AND (TRANSACTION_DATE BETWEEN '{}' AND '{}')".format(begin_date, end_date)

        # 1.3. get ORDER query
        order_query = " ORDER BY TRANSACTION_DATE;"

        # 2. Collect Data from SQL
        sql_query = select_query + where_query + order_query
        self.sqlBox.cursor.execute(sql_query)
        headers = ['DATE', 'BRANCH', 'IN', 'OUT', 'BALANCE', 'DESCRIPTION']
        count, balance = 0, 0
        total_in, total_out = 0, 0
        data_for_excel = []
        for row in self.sqlBox.cursor.fetchall():
            count += 1
            r_date, r_branch, r_cashflow, r_desc = row

            _in = Df.format_cost(r_cashflow) if r_cashflow > 0 else '-'
            _out = Df.format_cost(-r_cashflow) if r_cashflow <= 0 else '-'

            total_in += max(r_cashflow, 0)
            total_out += -min(r_cashflow, 0)

            balance = total_in - total_out
            new_row = (r_date, r_branch, _in, _out, Df.format_cost(balance), r_desc)
            data_for_excel.append(new_row)

        total_row = ('Total', '', Df.format_cost(total_in), Df.format_cost(total_out), Df.format_cost(balance), '')
        data_for_excel.append(total_row)

        # 3. Make Excel File
        try:
            df = pd.DataFrame(data_for_excel, columns=headers)
            wb = Workbook()
            ws = wb.active
            width_size = [12, 12, 12, 12, 12, 12]

            # add data frame onto Excel worksheet.
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
                for c_idx, value in enumerate(row):
                    cell = ws.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                    # styling headers
                    if r_idx == 0:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center")

                    # styling total row
                    if r_idx == len(df):
                        cell.font = Font(underline="single")
                        cell.border = Border(bottom=Side(style='thin'))

                    # modify cell's width
                    width_size[c_idx] = max(len(str(value)) * 1.5, width_size[c_idx])
                    column_width = width_size[c_idx]
                    ws.column_dimensions[cell.column_letter].width = column_width

            # save file
            time = str(datetime.now().timestamp()).replace('.', '_')
            excel_filename = 'daily_{}.xlsx'.format(time)
            wb.save(excel_filename)

            # Success Message
            print(f"Report successfully saved as {excel_filename}")
            print('...')
        except Exception as e:
            print('!Error:', e)
            print('...')
            return

        # 4. Print Summary
        print("\n*** Summary ***")
        print("- Branch: {}".format(self.branch.cur_path))
        print("- Period: {} ~ {}".format(begin_date, end_date))
        print("- Count: {}".format(count))
        print("- Total In: {}".format(Df.format_cost(total_in)))
        print("- Total Out: {}".format(Df.format_cost(total_out)))
        print("- Balance: {}".format(Df.format_cost(balance)))
        print()

    def excel_monthly(self, command: list):
        # 1. Get Path
        path = self.branch.cur_path

        # 2. get Query
        # 2.1. get SELECT query
        select_query = '''SELECT 
            STRFTIME('%Y-%m', TRANSACTION_DATE) AS MONTHLY, 
            SUM(CASE WHEN CASHFLOW > 0 THEN CASHFLOW ELSE 0 END) AS CASH_IN,
            SUM(CASE WHEN CASHFLOW < 0 THEN CASHFLOW ELSE 0 END) AS CASH_OUT
        FROM ACCOUNTBOOK'''

        # 2.2. get WHERE query
        where_query = "\n   WHERE BRANCH LIKE '{}%'".format(path)
        begin_date, end_date = '0001-01-01', '9999-12-31'
        if len(command) == 3:
            # check validity of period
            begin, end = command[2].split('~')
            if begin != '':
                begin_date = Df.format_date(begin)
            if end != '':
                end_date = Df.format_date(end)
            if (begin_date is None) or (end_date is None):
                print("!Error: NOT valid Date Input")
                print("...")
                return

        where_query += " AND (TRANSACTION_DATE BETWEEN '{}' AND '{}')".format(begin_date, end_date)

        # 2.3. get GROUP query
        group_query = " GROUP BY MONTHLY"

        # 2.4. get ORDER query
        order_query = " ORDER BY MONTHLY;"

        # 3. Execution on SQL
        sql_query = select_query + where_query + group_query + order_query
        self.sqlBox.cursor.execute(sql_query)

        # 4. Print table
        headers = ['MONTHLY', 'IN', 'OUT', 'BALANCE']
        data_for_excel = []
        total_in, total_out, balance = 0, 0, 0
        for row in self.sqlBox.cursor.fetchall():
            monthly, _in, _out = row
            total_in += _in
            total_out += -_out
            balance = (total_in - total_out)

            _in = Df.format_cost(_in) if _in != 0 else '-'
            _out = Df.format_cost(-_out) if _out != 0 else '-'

            new_row = (monthly, _in, _out, Df.format_cost(balance))
            data_for_excel.append(new_row)

        total_row = ('Total', Df.format_cost(total_in), Df.format_cost(total_out), Df.format_cost(balance))
        data_for_excel.append(total_row)

        # 5. Make Excel File
        try:
            df = pd.DataFrame(data_for_excel, columns=headers)
            wb = Workbook()
            ws = wb.active
            width_size = [12, 12, 12, 12]

            # add data frame onto Excel worksheet.
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
                for c_idx, value in enumerate(row):
                    cell = ws.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                    # styling headers
                    if r_idx == 0:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center")

                    # styling total row
                    if r_idx == len(df):
                        cell.font = Font(underline="single")
                        cell.border = Border(bottom=Side(style='thin'))

                    # modify cell's width
                    width_size[c_idx] = max(len(str(value)) * 1.5, width_size[c_idx])
                    column_width = width_size[c_idx]
                    ws.column_dimensions[cell.column_letter].width = column_width

            # save file
            time = str(datetime.now().timestamp()).replace('.', '_')
            excel_filename = 'monthly_{}.xlsx'.format(time)
            wb.save(excel_filename)

            # Success Message
            print(f"Report successfully saved as {excel_filename}")
            print('...')
        except Exception as e:
            print('!Error:', e)
            print('...')
            return

        print()
        print("*** Summary ***")
        print("- Branch: {}".format(self.branch.cur_path))
        print("- Period: {} ~ {}".format(begin_date, end_date))
        print("- Total In: {}".format(Df.format_cost(total_in)))
        print("- Total Out: {}".format(Df.format_cost(total_out)))
        print("- Balance: {}".format(Df.format_cost(balance)))
        print()

    def graph(self, command):
        if len(command) == 1:
            self.graph_balance(command)
        elif command[1] == 'in':
            self.graph_in(command)
        elif command[1] == 'out':
            self.graph_out(command)

    def graph_balance(self, command: list):
        # 1. Get Path
        path = self.branch.cur_path

        # 2. get Query
        # 2.1. get SELECT query
        select_query = '''SELECT
            STRFTIME('%Y-%m', TRANSACTION_DATE) AS MONTHLY, 
            SUM(CASE WHEN CASHFLOW > 0 THEN CASHFLOW ELSE 0 END) AS CASH_IN,
            SUM(CASE WHEN CASHFLOW < 0 THEN CASHFLOW ELSE 0 END) AS CASH_OUT
        FROM ACCOUNTBOOK'''

        # 2.2. get WHERE query
        where_query = "\n   WHERE BRANCH LIKE '{}%'".format(path)
        begin_date, end_date = '0001-01-01', '9999-12-31'
        if len(command) == 2:
            # check validity of period
            begin, end = command[2].split('~')
            if begin != '':
                begin_date = Df.format_date(begin)
            if end != '':
                end_date = Df.format_date(end)
            if (begin_date is None) or (end_date is None):
                print("!Error: NOT valid Date Input")
                print("...")
                return
        where_query += " AND (TRANSACTION_DATE BETWEEN '{}' AND '{}')".format(begin_date, end_date)

        # 2.3. get GROUP query
        group_query = " GROUP BY MONTHLY"

        # 2.4. get ORDER query
        order_query = " ORDER BY MONTHLY;"

        # 3. Get Data from SQL
        sql_query = select_query + where_query + group_query + order_query
        self.sqlBox.cursor.execute(sql_query)
        total_in, total_out, balance = 0, 0, 0
        months, balances = [], []
        for row in self.sqlBox.cursor.fetchall():
            monthly, _in, _out = row
            months.append(monthly)

            total_in += _in
            total_out += (-_out)
            balances.append(total_in - total_out)

        # 4. Print Summary
        print()
        print("*** Summary ***")
        print("- Branch: {}".format(self.branch.cur_path))
        print("- Period: {} ~ {}".format(begin_date, end_date))
        print("- Total In: {}".format(Df.format_cost(total_in)))
        print("- Total Out: {}".format(Df.format_cost(total_out)))
        print("- Balance: {}".format(Df.format_cost(total_in - total_out)))
        print()

        # 5. Bar graph
        plt.bar(months, balances, label='Balance')
        plt.xlabel('Month')
        plt.ylabel('Amount')
        plt.title('Monthly Balance')
        plt.legend()
        plt.show()

    def graph_in(self, command: list):
        # 1. Get Path
        path = self.branch.cur_path

        # 2. get Query
        # 2.1. get SELECT query
        select_query = '''SELECT 
            STRFTIME('%Y-%m', TRANSACTION_DATE) AS MONTHLY, 
            SUM(CASE WHEN CASHFLOW > 0 THEN CASHFLOW ELSE 0 END) AS CASH_IN
        FROM ACCOUNTBOOK'''

        # 2.2. get WHERE query
        where_query = "\n   WHERE BRANCH LIKE '{}%'".format(path)
        begin_date, end_date = '0001-01-01', '9999-12-31'
        if len(command) == 3:
            # check validity of period
            begin, end = command[2].split('~')
            if begin != '':
                begin_date = Df.format_date(begin)
            if end != '':
                end_date = Df.format_date(end)
            if (begin_date is None) or (end_date is None):
                print("!Error: NOT valid Date Input")
                print("...")
                return
        where_query += " AND (TRANSACTION_DATE BETWEEN '{}' AND '{}')".format(begin_date, end_date)

        # 2.3. get GROUP query
        group_query = " GROUP BY MONTHLY"

        # 2.4. get ORDER query
        order_query = " ORDER BY MONTHLY;"

        # 3. Get Data from SQL
        sql_query = select_query + where_query + group_query + order_query
        self.sqlBox.cursor.execute(sql_query)
        total_in = 0
        months = []
        cash_ins = []
        for row in self.sqlBox.cursor.fetchall():
            monthly, _in = row
            months.append(monthly)
            cash_ins.append(_in)
            total_in += _in

        # 4. Print Summary
        print()
        print("*** Summary ***")
        print("- Branch: {}".format(self.branch.cur_path))
        print("- Period: {} ~ {}".format(begin_date, end_date))
        print("- Total In: {}".format(Df.format_cost(total_in)))
        print()

        # 5. Bar graph..
        plt.bar(months, cash_ins, label='Cash In')
        plt.xlabel('Month')
        plt.ylabel('Amount')
        plt.title('Monthly Cash Flow')
        plt.legend()
        plt.show()

    def graph_out(self, command: list):
        # 1. Get Path
        path = self.branch.cur_path

        # 2. get Query
        # 2.1. get SELECT query
        select_query = '''SELECT 
            STRFTIME('%Y-%m', TRANSACTION_DATE) AS MONTHLY, 
            SUM(CASE WHEN CASHFLOW < 0 THEN CASHFLOW ELSE 0 END) AS CASH_IN
        FROM ACCOUNTBOOK'''

        # 2.2. get WHERE query
        where_query = "\n   WHERE BRANCH LIKE '{}%'".format(path)
        begin_date, end_date = '0001-01-01', '9999-12-31'
        if len(command) == 3:
            # check validity of period
            begin, end = command[2].split('~')
            if begin != '':
                begin_date = Df.format_date(begin)
            if end != '':
                end_date = Df.format_date(end)
            if (begin_date is None) or (end_date is None):
                print("!Error: NOT valid Date Input")
                print("...")
                return
        where_query += " AND (TRANSACTION_DATE BETWEEN '{}' AND '{}')".format(begin_date, end_date)

        # 2.3. get GROUP query
        group_query = " GROUP BY MONTHLY"

        # 2.4. get ORDER query
        order_query = " ORDER BY MONTHLY;"

        # 3. Get Data from SQL
        sql_query = select_query + where_query + group_query + order_query
        self.sqlBox.cursor.execute(sql_query)
        months, cash_outs = [], []
        total_out = 0
        for row in self.sqlBox.cursor.fetchall():
            monthly, _out = row
            months.append(monthly)
            cash_outs.append(_out)
            total_out += _out

        # 4. Print Summary
        print()
        print("*** Summary ***")
        print("- Branch: {}".format(self.branch.cur_path))
        print("- Period: {} ~ {}".format(begin_date, end_date))
        print("- Total Out: {}".format(Df.format_cost(total_out)))
        print()

        # 5. Bar Graph.
        plt.bar(months, cash_outs, label='Cash Out')
        plt.xlabel('Month')
        plt.ylabel('Amount')
        plt.title('Monthly Cash Flow')
        plt.legend()
        plt.show()

# pyinstaller --add-data "directory.json;." --icon=icon_financetree.ico -n FinanceTree3 main.py

# pyinstaller --add-data "directory.json;." --hidden-import=prettytable -n FinanceTree main.py

# pyinstaller --add-data "directory.json;." --hidden-import=prettytable --icon=icon_financetree.ico -n FinanceTree3 main.py
